from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def mark_rows_below_threshold(filename):
    # 打开Excel文件
    wb = load_workbook(filename)

    # 选择要操作的工作表
    ws = wb.active

    # 定义红色填充颜色
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # 遍历每一行，从第2行开始，假设第1行是标题行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # 获取该行中的分数列的值，假设分数列在第2列（B列）

        score_cell = row[2]  # 假设分数列在第2列(B列)，索引从0开始
        print(score_cell.value)
        # 检查分数是否小于等于5
        if score_cell.value is not None and int(score_cell.value) <= 5:
            # 如果分数小于等于5，将整行的填充颜色设置为红色
            for cell in row:
                cell.fill = red_fill

    # 保存修改后的Excel文件
    wb.save(filename)

